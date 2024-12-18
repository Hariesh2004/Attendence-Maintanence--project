import os
from openpyxl import Workbook, load_workbook
from datetime import datetime


FILE_PATH = "Attendance.xlsx"

def initialize_excel():
    
    if not os.path.exists(FILE_PATH):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        
        sheet.append(["ID", "Name", "Date", "Status"])
        workbook.save(FILE_PATH)
        print(f"Initialized new attendance file: {FILE_PATH}")

def mark_attendance():
    
    workbook = load_workbook(FILE_PATH)
    sheet = workbook.active

    
    print("\nMark Attendance:")
    id = input("Enter ID: ")
    name = input("Enter Name: ")
    date = input("Enter Date (YYYY-MM-DD) [leave blank for today]: ")
    status = input("Enter Status (Present/Absent): ")

    
    if not date:
        date = datetime.now().strftime("%Y-%m-%d")

    
    sheet.append([id, name, date, status])
    workbook.save(FILE_PATH)
    print("Attendance marked successfully!\n")

def view_attendance():
    """
    Display all attendance records.
    """
    
    workbook = load_workbook(FILE_PATH)
    sheet = workbook.active

    print("\nAttendance Records:")
    for row in sheet.iter_rows(values_only=True):
        print(row)

def main():
    
    
    initialize_excel()

    while True:
        print("\nAttendance System Menu:")
        print("1. Mark Attendance")
        print("2. View Attendance")
        print("3. Exit")

        choice = input("Enter your choice: ")

        if choice == "1":
            mark_attendance()
        elif choice == "2":
            view_attendance()
        elif choice == "3":
            print("Exiting the system. Goodbye!")
            break
        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
